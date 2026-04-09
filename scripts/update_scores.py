from pathlib import Path
import re
import json
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = REPO_ROOT / "data" / "2026 Masters Draft & Scoreboard.xlsx"
OUTPUT_JSON = REPO_ROOT / "docs" / "data" / "scores.json"

ESPN_URL = "https://www.espn.com/golf/leaderboard/_/pga"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

def normalize_name(name: str) -> str:
    name = name.replace("\xa0", " ").strip()
    name = re.sub(r"\s+", " ", name)
    return name

def fetch_html():
    response = requests.get(ESPN_URL, headers=HEADERS, timeout=30)
    response.raise_for_status()
    return response.text

def parse_scores(html: str):
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text("\n", strip=True)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    rows = []
    i = 0
    while i < len(lines):
        line = lines[i]

        if re.fullmatch(r"(T?\d+|CUT|WD|DQ|E)", line):
            chunk = lines[i:i+9]
            if len(chunk) >= 5:
                pos = chunk[0]
                player = normalize_name(chunk[1])
                score = chunk[2]
                today = chunk[3] if len(chunk) > 3 else ""
                thru = chunk[4] if len(chunk) > 4 else ""
                r1 = chunk[5] if len(chunk) > 5 else ""
                r2 = chunk[6] if len(chunk) > 6 else ""
                r3 = chunk[7] if len(chunk) > 7 else ""
                r4 = chunk[8] if len(chunk) > 8 else ""

                if len(player) > 2 and player.lower() != "player":
                    rows.append({
                        "pos": pos,
                        "player": player,
                        "score": score,
                        "today": today,
                        "thru": thru,
                        "r1": r1,
                        "r2": r2,
                        "r3": r3,
                        "r4": r4,
                    })
                    i += 9
                    continue
        i += 1

    if not rows:
        raise RuntimeError("Could not parse ESPN leaderboard. The page format may have changed.")

    return rows

def update_excel(scores):
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Scores"]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.value = None

    for row_index, item in enumerate(scores, start=2):
        ws.cell(row=row_index, column=1, value="")
        ws.cell(row=row_index, column=2, value=item["pos"])
        ws.cell(row=row_index, column=3, value=item["player"])
        ws.cell(row=row_index, column=4, value=item["score"])
        ws.cell(row=row_index, column=5, value=item["today"])
        ws.cell(row=row_index, column=6, value=item["thru"])
        ws.cell(row=row_index, column=7, value=item["r1"])
        ws.cell(row=row_index, column=8, value=item["r2"])
        ws.cell(row=row_index, column=9, value=item["r3"])
        ws.cell(row=row_index, column=10, value=item["r4"])

    wb.save(WORKBOOK_PATH)

def save_scores_json(scores):
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(scores, f, indent=2, ensure_ascii=False)

def main():
    html = fetch_html()
    scores = parse_scores(html)
    update_excel(scores)
    save_scores_json(scores)
    print(f"Updated Excel and JSON with {len(scores)} players.")

if __name__ == "__main__":
    main()