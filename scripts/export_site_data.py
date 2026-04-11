from pathlib import Path
import json
import re
import unicodedata
from datetime import datetime, timezone
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = REPO_ROOT / "data" / "2026 Masters Draft & Scoreboard.xlsx"
OUT_DIR = REPO_ROOT / "docs" / "data"

ALIASES = {
    "sam stevens": "samuel stevens",
    "nico echavarria": "nicolas echavarria",
    "johnny keefer": "john keefer",
}

def normalize_player_name(name):
    if name is None:
        return ""

    text = str(name).strip()
    text = re.sub(r"\s*\((a|A)\)\s*", "", text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    return ALIASES.get(text, text)

def normalize_score_display(pos, score, thru, today):
    pos_text = "" if pos is None else str(pos).strip().upper()
    score_text = "" if score is None else str(score).strip().upper()
    thru_text = "" if thru is None else str(thru).strip().upper()
    today_text = "" if today is None else str(today).strip().upper()

    if "CUT" in {pos_text, score_text, thru_text, today_text}:
        return "CUT"
    if "WD" in {pos_text, score_text, thru_text, today_text}:
        return "WD"
    if "DQ" in {pos_text, score_text, thru_text, today_text}:
        return "DQ"

    if pos_text in {"-", ""} and thru_text in {"", "-", "CUT"} and score_text in {"", "-", "CUT"}:
        return "CUT"

    return "" if score is None else str(score).strip()

def score_to_number(value):
    if value is None or value == "":
        return None

    text = str(value).strip().upper()

    if text in {"E", "(E)"}:
        return 0
    if text in {"CUT", "WD", "DQ"}:
        return None

    text = text.replace("(", "").replace(")", "")
    try:
        return int(text)
    except ValueError:
        return None

def round_score_to_number(value):
    if value is None or value == "":
        return None
    text = str(value).strip().upper()
    if text in {"--", "-", "CUT", "WD", "DQ"}:
        return None
    try:
        return int(text)
    except ValueError:
        return None

def is_real_score_row(pos, player, score, thru):
    if not player:
        return False

    player = str(player).strip()
    pos = str(pos).strip() if pos is not None else ""
    score = str(score).strip().upper() if score is not None else ""
    thru = str(thru).strip().upper() if thru is not None else ""

    banned_players = {"PLAYER", "YARDS", "TOURNAMENTS", "PREVIOUS WINNER", "HIDDEN"}
    if player.upper() in banned_players:
        return False

    if player.isdigit() or any(ch.isdigit() for ch in player):
        return False

    if pos and not re.fullmatch(r"(T?\d+|CUT|WD|DQ|-)", pos):
        return False

    if score and not re.fullmatch(r"(E|[+-]?\d+|CUT|WD|DQ|-)", score):
        return False

    if thru and not re.fullmatch(r"(\d+|F|CUT|WD|DQ|-|[0-9]{1,2}:[0-9]{2}\s?[AP]M)", thru):
        return False

    return True

def split_payout(label, winners, total_amount):
    winners = sorted(set(winners))
    if not winners:
        return None

    share = round(total_amount / len(winners), 2)
    return {
        "label": label,
        "winners": [{"name": winner, "amount": share} for winner in winners]
    }

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    scores_ws = wb["Scores"]
    draft_ws = wb["Draft_Import"]

    scores_lookup = {}
    score_rows = []
    player_to_teams = defaultdict(set)

    for row in draft_ws.iter_rows(min_row=2, max_row=draft_ws.max_row, min_col=1, max_col=5, values_only=True):
        team = row[3]
        player = row[4]
        if not team or not player:
            continue

        team_name = str(team).strip()
        player_name = normalize_player_name(player)
        player_to_teams[player_name].add(team_name)

    for row in scores_ws.iter_rows(min_row=2, max_row=scores_ws.max_row, min_col=1, max_col=11, values_only=True):
        pos = row[1]
        player = row[3]
        score = row[4]
        today = row[5]
        thru = row[6]

        if not is_real_score_row(pos, player, score, thru):
            continue

        lookup_name = normalize_player_name(player)
        display_score = normalize_score_display(pos, score, thru, today)
        numeric_score = score_to_number(display_score)

        r1 = round_score_to_number(row[7])
        r2 = round_score_to_number(row[8])
        r3 = round_score_to_number(row[9])

        day3_total = None
        if r1 is not None and r2 is not None and r3 is not None:
            day3_total = r1 + r2 + r3

        entry = {
            "pos": "" if pos is None else str(pos).strip(),
            "player": str(player).strip(),
            "score": display_score,
            "today": "" if today is None else str(today).strip(),
            "thru": "" if thru is None else str(thru).strip(),
            "r1": "" if row[7] is None else str(row[7]).strip(),
            "r2": "" if row[8] is None else str(row[8]).strip(),
            "r3": "" if row[9] is None else str(row[9]).strip(),
            "r4": "" if row[10] is None else str(row[10]).strip(),
            "numeric_score": numeric_score,
            "day3_total": day3_total,
        }

        scores_lookup[lookup_name] = entry
        score_rows.append(entry)

    teams_map = OrderedDict()
    for team_name in sorted({t for teams in player_to_teams.values() for t in teams}):
        teams_map[team_name] = []

    for player_name, teams_for_player in player_to_teams.items():
        for team_name in teams_for_player:
            teams_map[team_name].append(player_name)

    teams = []
    for team_name, golfers in teams_map.items():
        golfer_details = []
        valid_scores = []
        valid_day3_scores = []

        for golfer in golfers:
            info = scores_lookup.get(golfer, {
                "pos": "",
                "player": golfer,
                "score": "",
                "today": "",
                "thru": "",
                "r1": "",
                "r2": "",
                "r3": "",
                "r4": "",
                "numeric_score": None,
                "day3_total": None,
            })

            golfer_details.append(info)

            if info["numeric_score"] is not None:
                valid_scores.append(info["numeric_score"])
            if info["day3_total"] is not None:
                valid_day3_scores.append(info["day3_total"])

        best_three = sorted(valid_scores)[:3]
        best_three_total = sum(best_three) if len(best_three) >= 3 else None

        best_three_day3 = sorted(valid_day3_scores)[:3]
        best_three_day3_total = sum(best_three_day3) if len(best_three_day3) >= 3 else None

        teams.append({
            "team": team_name,
            "golfers": golfer_details,
            "best3_total": best_three_total,
            "best3_day3_total": best_three_day3_total,
            "scores_entered": len(valid_scores),
            "roster_loaded": len(golfers),
        })

    teams_sorted = sorted(
        teams,
        key=lambda t: (99999 if t["best3_total"] is None else t["best3_total"], t["team"])
    )

    payouts = []

    valid_day3_players = [s for s in score_rows if s["day3_total"] is not None]
    if valid_day3_players:
        best_day3_score = min(p["day3_total"] for p in valid_day3_players)
        winning_players = [p for p in valid_day3_players if p["day3_total"] == best_day3_score]

        winning_teams = []
        for p in winning_players:
            lookup_name = normalize_player_name(p["player"])
            winning_teams.extend(player_to_teams.get(lookup_name, []))

        item = split_payout("Leader after Day 3", winning_teams, 50)
        if item:
            payouts.append(item)

    valid_day3_teams = [t for t in teams if t["best3_day3_total"] is not None]
    if valid_day3_teams:
        best_team_score = min(t["best3_day3_total"] for t in valid_day3_teams)
        winners = [t["team"] for t in valid_day3_teams if t["best3_day3_total"] == best_team_score]

        item = split_payout("Best 3-Man Team after Day 3", winners, 50)
        if item:
            payouts.append(item)

    with open(OUT_DIR / "teams.json", "w", encoding="utf-8") as f:
        json.dump(teams_sorted, f, indent=2, ensure_ascii=False)

    with open(OUT_DIR / "scores.json", "w", encoding="utf-8") as f:
        json.dump(score_rows, f, indent=2, ensure_ascii=False)

    with open(OUT_DIR / "payouts.json", "w", encoding="utf-8") as f:
        json.dump({"items": payouts}, f, indent=2, ensure_ascii=False)

    with open(OUT_DIR / "meta.json", "w", encoding="utf-8") as f:
        json.dump({"last_updated": datetime.now(timezone.utc).isoformat()}, f, indent=2)

    print("Exported website data with Day 3 payouts.")

if __name__ == "__main__":
    main()
